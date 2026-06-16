"""
EIA capacity, planned additions, and retirements data.

Supports O_j (Market Mismatch / Operational Fragility Risk) in the
AI infrastructure risk framework.
Indicators: demand uncertainty, grid capacity vs. projected data center load,
planned additions, reliance on new capacity, stranded-asset risk.
Cited in paper as EIAElectricityData and epri2024ai.

Downloads:
  1. EIA-860M (monthly update) — planned generating capacity additions and retirements
  2. EIA Electric Power Monthly — state retail electricity prices and sales
  3. EIA Annual Energy Outlook state-level forecast data (where available)
"""

import csv
import io
import json
import urllib.request
import zipfile
from pathlib import Path

OUT_DIR = Path(__file__).parent

NE_STATE = "NE"
NEBRASKA_FIPS = "31"

# EIA-860M: monthly update on planned additions and retirements
# URL pattern: https://www.eia.gov/electricity/data/eia860m/xls/{month}_generator{year}.xlsx
EIA860M_URLS = [
    # EIA-860M: try archive paths which are more stable
    "https://www.eia.gov/electricity/data/eia860m/archive/xls/december_generator2023.xlsx",
    "https://www.eia.gov/electricity/data/eia860m/archive/xls/september_generator2023.xlsx",
    "https://www.eia.gov/electricity/data/eia860m/archive/xls/june_generator2023.xlsx",
]
# Fallback: EIA-860 annual archive (2022) for capacity data
EIA860_ARCHIVE_URL = "https://www.eia.gov/electricity/data/eia860/archive/xls/eia8602022.zip"

# EIA Electric Power Monthly — Table 5.6: Average Retail Price of Electricity by State
EPM_STATE_PRICE_URLS = [
    "https://www.eia.gov/electricity/data/state/sales_revenue.zip",
    "https://www.eia.gov/electricity/monthly/archive/december2024.pdf",  # for reference
]

# EIA average retail prices by state (simpler direct download)
EIA_RETAIL_URLS = [
    "https://www.eia.gov/electricity/data/state/avgprice_annual.xls",
]


def download_file(url: str, out_path: Path) -> bool:
    print(f"  Downloading {out_path.name} ...")
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=120) as r:
            data = r.read()
        with open(out_path, "wb") as f:
            f.write(data)
        print(f"    Saved {len(data) / 1e6:.1f} MB")
        return True
    except Exception as e:
        print(f"    Failed: {e}")
        return False


def try_urls(urls: list[str], out_path: Path) -> bool:
    for url in urls:
        if download_file(url, out_path):
            return True
    return False


def extract_nebraska_from_860m(path: Path) -> None:
    """Extract Nebraska planned additions and retirements from EIA-860M."""
    try:
        import openpyxl
    except ImportError:
        print("  openpyxl not installed; skipping EIA-860M extraction")
        return

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        print(f"  EIA-860M sheets: {wb.sheetnames}")

        for sheet_name in wb.sheetnames:
            # EIA-860M sheets: "Operating", "Planned", "Retired", "Canceled or Postponed", etc.
            if sheet_name.lower() not in {"operating", "planned", "retired", "canceled or postponed"}:
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 3:
                continue

            # Find header row
            hdr_idx = next(
                (i for i, r in enumerate(rows[:5]) if sum(1 for c in r if c is not None and str(c).strip()) > 4),
                0
            )
            headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[hdr_idx])]
            data_rows = rows[hdr_idx + 1:]

            state_col = next(
                (i for i, h in enumerate(headers) if "state" in h.lower()),
                None
            )
            if state_col is None:
                continue

            ne_rows = [r for r in data_rows if str(r[state_col] or "").upper() == NE_STATE]
            if ne_rows:
                ne_path = OUT_DIR / f"eia860m_{sheet_name.lower().replace(' ', '_')}_nebraska.csv"
                with open(ne_path, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)
                    writer.writerows(ne_rows)
                print(f"  Nebraska '{sheet_name}': {len(ne_rows)} rows -> {ne_path.name}")

        wb.close()
    except Exception as e:
        print(f"  EIA-860M extraction error: {e}")


def read_excel_to_rows(path: Path) -> list:
    """Read an xls or xlsx file into a list of row-tuples."""
    if path.suffix.lower() == ".xls":
        import xlrd
        wb = xlrd.open_workbook(str(path))
        ws = wb.sheet_by_index(0)
        return [ws.row_values(i) for i in range(ws.nrows)]
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def extract_nebraska_from_retail_prices(path: Path) -> None:
    """Extract Nebraska electricity retail prices from EIA state price file."""
    try:
        rows = read_excel_to_rows(path)

        hdr_idx = next(
            (i for i, r in enumerate(rows[:5]) if sum(1 for c in r if c is not None and str(c).strip()) > 4),
            0
        )
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[hdr_idx])]
        data_rows = rows[hdr_idx + 1:]

        state_col = next(
            (i for i, h in enumerate(headers) if "state" in str(h).lower()),
            0
        )

        ne_rows = [r for r in data_rows if str(r[state_col] or "").upper().strip() == NE_STATE]

        nat_path = OUT_DIR / "avgprice_annual_national.csv"
        with open(nat_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(data_rows)
        print(f"  Retail prices national: {nat_path.name} ({len(data_rows)} rows)")

        if ne_rows:
            ne_path = OUT_DIR / "avgprice_annual_nebraska.csv"
            with open(ne_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                writer.writerows(ne_rows)
            print(f"  Nebraska retail prices: {len(ne_rows)} rows -> {ne_path.name}")

    except Exception as e:
        print(f"  Retail price extraction error: {e}")


def write_risk_framework_note() -> None:
    """Write a JSON note connecting this data to the O_j risk variable."""
    note = {
        "risk_variable": "O_j (Market Mismatch / Operational Fragility)",
        "definition": (
            "Captures downside risks to the operational and financial viability "
            "of AI infrastructure investments."
        ),
        "key_indicators": {
            "planned_additions": (
                "New generation capacity planned — indicates whether grid can absorb "
                "large incremental AI data center demand without stranding new assets"
            ),
            "planned_retirements": (
                "Retiring generation — reduces reserve margins, may force new procurement "
                "under unfavorable conditions"
            ),
            "retail_electricity_price": (
                "Average price (cents/kWh) for commercial/industrial customers — "
                "directly affects data center operating cost and community rate impacts"
            ),
            "demand_uncertainty": (
                "Gap between projected AI compute demand and realized utilization — "
                "not directly downloadable; requires analyst judgment from EPRI/company disclosures"
            ),
        },
        "nebraska_relevance": (
            "Nebraska's publicly owned utilities (OPPD, LES, NPPD) operate with limited "
            "reserve margins. Large incremental loads from hyperscale data centers "
            "increase E_marginal and risk of cost overruns, stranded assets, or rate increases."
        ),
        "additional_sources": {
            "EPRI_AI_report": "https://www.epri.com/research/products/000000003002028905",
            "NERC_reliability": "https://www.nerc.com/pa/RAPA/ra/Pages/default.aspx",
            "EIA_electric_power_monthly": "https://www.eia.gov/electricity/monthly/",
        },
    }
    note_path = OUT_DIR / "market_mismatch_data_note.json"
    with open(note_path, "w", encoding="utf-8") as f:
        json.dump(note, f, indent=2)
    print(f"  Risk framework note: {note_path.name}")


def main():
    print("Downloading EIA capacity and retail price data for market mismatch risk...\n")

    # EIA-860M planned additions and retirements
    print("1. EIA-860M Monthly Capacity Update (planned additions/retirements)...")
    eia860m_path = OUT_DIR / "eia860m_latest.xlsx"
    if try_urls(EIA860M_URLS, eia860m_path):
        extract_nebraska_from_860m(eia860m_path)
    else:
        print("  EIA-860M download failed.")
        print("  Visit: https://www.eia.gov/electricity/data/eia860m/")
    print()

    # Average retail electricity prices
    print("2. EIA Average Retail Electricity Prices by State...")
    retail_path = OUT_DIR / "avgprice_annual.xls"
    if try_urls(EIA_RETAIL_URLS, retail_path):
        extract_nebraska_from_retail_prices(retail_path)
    else:
        print("  Retail price download failed.")
        print("  Visit: https://www.eia.gov/electricity/data/state/")
    print()

    write_risk_framework_note()
    print("\nMarket mismatch data download complete.")


if __name__ == "__main__":
    main()
