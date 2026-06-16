"""
EPA eGRID 2022 — power plant emissions, generation, and water use by subregion.

Supports R_res_j / E_j (Electricity Risk) and W_j (indirect water via grid) in the
AI infrastructure risk framework. Cited in paper as epa2024egrid.

Nebraska is served by MRO West (MROW) and SPP North (SPPN) / SPP South (SPPS) subregions.
Key metrics: emission rates, generation mix, water withdrawal/consumption intensities.
"""

import csv
import urllib.request
from pathlib import Path

OUT_DIR = Path(__file__).parent

EGRID_URL = "https://www.epa.gov/system/files/documents/2024-01/egrid2022_data.xlsx"

# eGRID subregions that serve Nebraska
NE_SUBREGIONS = {"MROW", "SPPN", "SPPS"}
NE_STATE_ABBR = "NE"


def download_egrid(url: str, out_path: Path) -> bool:
    print(f"  Downloading eGRID from {url} ...")
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    try:
        with urllib.request.urlopen(req, timeout=180) as r:
            data = r.read()
        with open(out_path, "wb") as f:
            f.write(data)
        print(f"  Saved {len(data) / 1e6:.1f} MB -> {out_path.name}")
        return True
    except Exception as e:
        print(f"  Download failed: {e}")
        return False


def extract_sheets(xlsx_path: Path) -> None:
    try:
        import openpyxl
    except ImportError:
        print("  openpyxl not installed. Run: pip install openpyxl")
        return

    print("  Extracting sheets from eGRID Excel workbook...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    print(f"  Available sheets: {wb.sheetnames}")

    # eGRID 2022 sheet naming uses year suffix (PLNT22, ST22, SRL22, etc.)
    # Map each known sheet name pattern -> descriptive label
    raw_sheets = wb.sheetnames
    # Build mapping: try year-suffixed names, then legacy names
    candidate_map = {
        "SRL22": "subregion",   # subregion-level emission rates, generation mix, water use
        "SUBRGN": "subregion",
        "PLNT22": "plant",       # plant-level data
        "PLNT": "plant",
        "ST22": "state",         # state-level
        "ST": "state",
        "US22": "us_total",
        "US": "us_total",
    }
    target_sheets = {}
    for sheet_name, label in candidate_map.items():
        if sheet_name in raw_sheets and label not in target_sheets.values():
            target_sheets[sheet_name] = label

    for sheet_name, label in target_sheets.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # eGRID 2022 workbook structure:
        #   Row 0: long-form column descriptions
        #   Row 1: short variable codes (PSTATABB, SUBRGN, etc.)
        #   Row 2+: data
        # Use row 1 (short codes) as headers so filtering by PSTATABB works.
        if len(rows) < 3:
            continue
        header_row_idx = 1  # short variable codes
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[header_row_idx])]
        data_rows = rows[header_row_idx + 1:]

        # Save full national CSV
        nat_path = OUT_DIR / f"egrid2022_{label}_national.csv"
        with open(nat_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(data_rows)
        print(f"  {sheet_name}: {len(data_rows)} rows -> {nat_path.name}")

        # Filter Nebraska-relevant rows
        ne_rows = []
        for row in data_rows:
            row_dict = dict(zip(headers, row))
            # Match by state abbreviation (plant/state sheets)
            state_val = str(row_dict.get("PSTATABB", row_dict.get("STABB", ""))).upper()
            # Match by subregion code (try multiple column name variants)
            subrgn_val = str(
                row_dict.get("SUBRGN", row_dict.get("SRNAME", row_dict.get("SRL22SUBRGN", "")))
            ).upper()

            if state_val == NE_STATE_ABBR or any(s in subrgn_val for s in NE_SUBREGIONS):
                ne_rows.append(row)

        if ne_rows:
            ne_path = OUT_DIR / f"egrid2022_{label}_nebraska_relevant.csv"
            with open(ne_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                writer.writerows(ne_rows)
            print(f"  Nebraska-relevant {sheet_name}: {len(ne_rows)} rows -> {ne_path.name}")

    wb.close()


def main():
    print("Downloading EPA eGRID 2022...")
    xlsx_path = OUT_DIR / "egrid2022_data.xlsx"

    if download_egrid(EGRID_URL, xlsx_path):
        extract_sheets(xlsx_path)
    else:
        print("\n  Manual download instructions:")
        print("  1. Visit: https://www.epa.gov/egrid/download-data")
        print("  2. Download eGRID2022 Excel file")
        print("  3. Place egrid2022_data.xlsx in this directory and re-run")


if __name__ == "__main__":
    main()
