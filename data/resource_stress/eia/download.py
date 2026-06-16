"""
EIA electricity data — state-level generation, capacity, and plant-level data.

Supports R_res_j / E_j (Electricity Risk) in the AI infrastructure risk framework.
Indicators: total electricity load (E_load), grid capacity margins (E_marginal),
grid water intensity (S_grid), generation mix. Cited in paper as EIAElectricityData.

Downloads:
  - Annual state-level electricity generation (all states)
  - Existing capacity by state
  - EIA Form 860 (plant-level data, generation technology, Nebraska filter)
"""

import csv
import io
import urllib.request
import zipfile
from pathlib import Path

OUT_DIR = Path(__file__).parent

NE_STATE = "NE"

DATASETS = {
    "annual_generation_state.xls": "https://www.eia.gov/electricity/data/state/annual_generation_state.xls",
    "existcapacity_annual.xlsx": "https://www.eia.gov/electricity/data/state/existcapacity_annual.xlsx",
}

# EIA-860 annual: plant-level capacity, fuel type, ownership
EIA860_URLS = [
    # 2022 archive works; 2023 URL currently serves HTML
    "https://www.eia.gov/electricity/data/eia860/archive/xls/eia8602022.zip",
    "https://www.eia.gov/electricity/data/eia860/archive/xls/eia8602021.zip",
]

# EIA-923 monthly: plant-level generation and fuel consumption
EIA923_URLS = [
    "https://www.eia.gov/electricity/data/eia923/archive/xls/f923_2022.zip",
    "https://www.eia.gov/electricity/data/eia923/archive/xls/f923_2021.zip",
]


def download_file(url: str, out_path: Path) -> bool:
    print(f"  Downloading {out_path.name} ...")
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=180) as r:
            data = r.read()
        with open(out_path, "wb") as f:
            f.write(data)
        print(f"    Saved {len(data) / 1e6:.1f} MB")
        return True
    except Exception as e:
        print(f"    Failed: {e}")
        return False


def try_urls(urls: list[str], out_path: Path) -> bool:
    for url in urls:
        if download_file(url, out_path):
            return True
    return False


def read_excel_rows(path: Path):
    """Read all rows from an Excel file (.xls or .xlsx). Returns (headers, data_rows)."""
    suffix = path.suffix.lower()
    if suffix == ".xls":
        import xlrd
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)
        all_rows = [ws.row_values(i) for i in range(ws.nrows)]
    else:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
    return all_rows


def extract_nebraska_xls(path: Path, label: str) -> None:
    """Extract Nebraska rows from an EIA xls/xlsx file."""
    try:
        all_rows = read_excel_rows(path)
        if len(all_rows) < 2:
            return

        # Find header row (first row with >3 non-empty cells)
        hdr_idx = next(
            (i for i, row in enumerate(all_rows[:5]) if sum(1 for c in row if c) > 3),
            0
        )
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(all_rows[hdr_idx])]
        data_rows = all_rows[hdr_idx + 1:]

        # Detect state column
        state_col_idx = next(
            (i for i, h in enumerate(headers) if h.strip().lower() in {"state", "stateid", "st"}),
            None
        )
        if state_col_idx is None:
            # Broader search
            state_col_idx = next(
                (i for i, h in enumerate(headers) if "state" in h.lower()),
                None
            )
        if state_col_idx is None:
            print(f"    No state column found in {path.name}")
            return

        ne_rows = [r for r in data_rows if str(r[state_col_idx] or "").upper().strip() == NE_STATE]
        if ne_rows:
            ne_path = OUT_DIR / f"{label}_nebraska.csv"
            with open(ne_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                writer.writerows(ne_rows)
            print(f"    Nebraska rows: {len(ne_rows)} -> {ne_path.name}")

        # National CSV
        national_path = OUT_DIR / f"{label}_national.csv"
        with open(national_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(data_rows)
        print(f"    National CSV: {national_path.name} ({len(data_rows)} rows)")

    except Exception as e:
        print(f"    Extraction error: {e}")


def extract_nebraska_from_eia860_zip(zip_path: Path) -> None:
    """Extract Nebraska plant data from EIA-860 zip."""
    try:
        import openpyxl
    except ImportError:
        print("    openpyxl not installed; skipping EIA-860 extraction")
        return

    try:
        with zipfile.ZipFile(zip_path) as z:
            xlsx_files = [n for n in z.namelist() if n.endswith(".xlsx")]
            print(f"    EIA-860 xlsx files: {xlsx_files}")

            # "2___Plant_Y20XX.xlsx" has plant-level location and ownership data
            plant_files = [n for n in xlsx_files if "Plant" in n or "plant" in n]
            generator_files = [n for n in xlsx_files if "Generator" in n or "generator" in n]

            for xlsx_name in plant_files[:1] + generator_files[:1]:
                label = "eia860_plant" if "Plant" in xlsx_name else "eia860_generator"
                with z.open(xlsx_name) as f:
                    wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True, read_only=True)

                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = list(ws.iter_rows(values_only=True))
                    if len(rows) < 3:
                        continue

                    # EIA-860 Plant sheet: row 0 = title/description, row 1 = real headers, row 2+ = data
                    # Find the header row by looking for "State" in one of the first few rows
                    hdr_idx = next(
                        (i for i, r in enumerate(rows[:4]) if any(str(c or "").strip() == "State" for c in r)),
                        1
                    )
                    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[hdr_idx])]
                    data_rows = rows[hdr_idx + 1:]

                    state_col = next(
                        (i for i, h in enumerate(headers) if str(h).strip() == "State"),
                        None
                    )
                    if state_col is None:
                        state_col = next(
                            (i for i, h in enumerate(headers) if "state" in str(h).lower()),
                            None
                        )
                    if state_col is None:
                        continue

                    ne_rows = [r for r in data_rows if str(r[state_col] or "").upper() == NE_STATE]
                    if ne_rows:
                        ne_path = OUT_DIR / f"{label}_nebraska.csv"
                        with open(ne_path, "w", newline="", encoding="utf-8") as f:
                            writer = csv.writer(f)
                            writer.writerow(headers)
                            writer.writerows(ne_rows)
                        print(f"    Nebraska {label}: {len(ne_rows)} rows -> {ne_path.name}")
                    wb.close()
                    break

    except Exception as e:
        print(f"    EIA-860 extraction error: {e}")


def main():
    print("Downloading EIA electricity datasets...\n")

    # State-level generation and capacity
    for filename, url in DATASETS.items():
        out_path = OUT_DIR / filename
        if download_file(url, out_path):
            label = filename.replace(".xls", "").replace(".xlsx", "")
            extract_nebraska_xls(out_path, label)
        print()

    # EIA Form 860 — plant-level capacity data
    print("Downloading EIA Form EIA-860 (plant capacity data)...")
    eia860_path = OUT_DIR / "eia860.zip"
    if try_urls(EIA860_URLS, eia860_path):
        extract_nebraska_from_eia860_zip(eia860_path)
    else:
        print("  EIA-860 download failed. Visit: https://www.eia.gov/electricity/data/eia860/")
    print()

    # EIA Form 923 — monthly generation and fuel consumption
    print("Downloading EIA Form EIA-923 (monthly generation data)...")
    eia923_path = OUT_DIR / "eia923.zip"
    if try_urls(EIA923_URLS, eia923_path):
        print("  EIA-923 downloaded. Use openpyxl to extract generation by plant.")
    else:
        print("  EIA-923 download failed. Visit: https://www.eia.gov/electricity/data/eia923/")

    print("\nEIA download complete.")


if __name__ == "__main__":
    main()
